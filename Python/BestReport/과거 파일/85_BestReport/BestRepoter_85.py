import tkinter
from openpyxl import Workbook, load_workbook
from tkinter.ttk import LabelFrame
from tkinter import filedialog


#######윈도우 생성#######
window=tkinter.Tk()
window.title("Best Reporter")
window.geometry("400x300+800+50")
window.resizable(False,False)

#######동작 함수#######
def fileOpen():
    #UI 화면에 진행 상황 출력 시키기
    log_text.config(text="시작")
    #파일 탐색기 오픈해서 선택한 파일의 경로를 filename에 저장  
    filename=filedialog.askopenfilename(initialdir="",title="Select File")
    #선택한 파일 
    log_text.config(text=filename)
    
    #######엑셀 변수 생성#######    row =1,2,3... // column=a,b,c.....
    newWb=Workbook()            #새로운 워크북 생성(우테 취합용)
    wb=load_workbook(filename)  #버튼으로 선택한 엑셀 파일 불러오기

    #버그리포트 취합 엑셀 전체 시트 중 뒤에서 1번째(우테 취합)까지 이름 가져오기    
    for ws,i in zip(wb.sheetnames, range(1,len(wb.sheetnames))): 

        brws = newWb.create_sheet(ws)
        print(ws)   #워크시트 이름 하나씩 출력
        log_text.config(text=ws)
        count=2
        brws.cell(row=1,column=2).value="JIRA 링크"
        brws.cell(row=1,column=7).value="처리 여부"
        brws.cell(row=1,column=8).value="등급"
        brws.cell(row=1,column=9).value="등록자"
        brws.cell(row=1,column=11).value="글번호"
        brws.cell(row=1,column=12).value="담당자"

        for j in wb[ws].rows:        #해당 워크 시트 중 열의 끝까지 반복
            if (j[5].value == "JIRA 등록"): #해당 워크 시트의 5행의 열이 "JIRA 등록일 때"
                #print(j[1].value)
                #print(j[7].value)  #참일 경우 1열의 "글번호" 7열의 "JIRA 링크" 출력
                brws.cell(row=count,column=2).value=j[7].value  #JIRA 링크
                brws.cell(row=count,column=7).value=j[5].value  #처리 여부
                brws.cell(row=count,column=8).value=j[6].value  #등급
                brws.cell(row=count,column=9).value=j[2].value  #등록자
                brws.cell(row=count,column=11).value=j[1].value  #글번호
                brws.cell(row=count,column=12).value=j[4].value  #담당자
                count=count+1
            elif (j[5].value =="본섭수정"):
                #print(j[1].value)
                #print(j[3].value)   #글번호와 내용 부분 출력
                brws.cell(row=count,column=2).value=j[3].value  #글 요약
                brws.cell(row=count,column=7).value=j[5].value  #처리 여부
                brws.cell(row=count,column=8).value="B"  #등급
                brws.cell(row=count,column=9).value=j[2].value  #등록자
                brws.cell(row=count,column=11).value=j[1].value  #글번호
                brws.cell(row=count,column=12).value=j[4].value  #담당자
                count=count+1
            elif(j[5].value=="NextUpdate"):
                #print(j[1].value)
                #print(j[7].value)
                brws.cell(row=count,column=2).value=j[7].value  #JIRA 링크
                brws.cell(row=count,column=7).value=j[5].value  #처리 여부
                brws.cell(row=count,column=8).value="B"  #등급
                brws.cell(row=count,column=9).value=j[2].value  #등록자
                brws.cell(row=count,column=11).value=j[1].value  #글번호
                brws.cell(row=count,column=12).value=j[4].value  #담당자
                count=count+1

        print("---")

            
    log_text.config(text="끝")
    newWb.save("BestTester.xlsx")

def fileEdit():
    filename=filedialog.askopenfilename(initialdir="",title="Select File")
    #선택한 파일 
    log_text.config(text=filename)
    
    filewb=load_workbook(filename)  #버튼으로 선택한 엑셀 파일 불러오기
    print(filewb.sheetnames)
    for ws, i in zip(filewb.sheetnames, range(1,len(filewb.sheetnames))):
        if (ws == "Sheet"):
            #jira 추출한거 넣어주기
            print("jira 추출 값 넣어주기")
            #filewb[ws].cell(row=1,column=1).value = "jira 추출 값 넣어주기"
        else:
            count=2
            for r in filewb[ws].rows:               #워크시트의 열의 갯수만큼 반복
                ####함수 넣기#####
                if (filewb[ws].cell(row=count, column=7).value=="JIRA 등록"):     #워크시트의 처리 여부가 jira 등록일 경우 
                    filewb[ws].cell(row=count,column=3).value=f"=VLOOKUP(B{count},Sheet!B:C,2,0)"   # jira 제목 찾기

                elif (filewb[ws].cell(row=count, column=7).value=="NextUpdate"):  #워크시트의 처리 여부가 넥업일 경우
                    filewb[ws].cell(row=count,column=3).value=f"=VLOOKUP(B{count},Sheet!B:C,2,0)"   # jira 제목 찾기

                else:                   # 이외의 상황에서는 내용의 값을 찾아오기
                    filewb[ws].cell(row=count, column=3).value=filewb[ws].cell(row=count, column=2).value
                
                ####JIRA 링크 문자열 편집#####
                if(filewb[ws].cell(row=count, column=2).value != None):     #각 워크 시트의 B열의 값이 있다면
                    jiraLink=filewb[ws].cell(row=count, column=2).value.strip() #공백 삭제
                    if (jiraLink.find("M") != -1):                      #문자열에 M이 있을 경우
                        jiraLink_index=jiraLink.index("M")                  #문자열 중 M이 있는 인덱스 값 찾기
                        jiraLink_Edit=jiraLink[jiraLink_index:jiraLink_index+11]    #JIRA 번호 추출
                        #print(jiraLink_Edit)
                        filewb[ws].cell(row=count, column=2).value=jiraLink_Edit      #셀에 작성해주기
                        count+=1
                    else:               # 문자열에 M을 못 찾았거나 오류 발생 시
                        #print(jiraLink.find("M"))
                        filewb[ws].cell(row=count, column=2).value = "Jira 링크 없음"
                        count+=1
    #새로운 탭 만들고 등록자와 jira 등급 가져오기
    filewb.create_sheet("등록자 카운팅") #새로운 워크 시트 생성
    filewb['등록자 카운팅'].cell(row=1, column=2).value="JIRA 등급" # 제목 입력
    filewb['등록자 카운팅'].cell(row=1, column=3).value="등록자"    # 제목 입력
    total_count=2
    for fws,sc in zip(filewb.sheetnames, range(1,len(filewb.sheetnames))):  #1~워크시트 리스트 개수만큼 반복, fws = 개별 워크시크 이릅, sc=시트번호
        r_count=2
        if(sc >= 2):        #시트 넘버가 2번 이상일 경우
            for r in filewb[fws].rows:      #해당 워크시트의 열 개수만큼 반복
                if (filewb[fws].cell(row=r_count, column=8).value != None): # h행의 값이 빈 값이 아닐 경우 
                    filewb['등록자 카운팅'].cell(row=total_count, column=2).value = filewb[fws].cell(row=r_count, column=8).value  #새로운 워크 시트의 b행의 2열부터 값 입력(jira 등급)
                    reporterID=str(filewb[fws].cell(row=r_count, column=9).value).strip() #새로운 워크 시트의 c행의 2열부터 값 입력 (등록자)
                    filewb['등록자 카운팅'].cell(row=total_count, column=3).value = reporterID
                    r_count+=1
                    total_count+=1
                    print(total_count)

    filewb['등록자 카운팅'].cell(row=1, column=7).value="등록자 중복 제거"
    l_count=2
    reporterList = []
    for l in filewb['등록자 카운팅'].rows:
        reporterList.append(str(filewb['등록자 카운팅'].cell(row=l_count, column=3).value))
        l_count+=1
    print(reporterList)

    log_text.config(text="끝")
    filewb.save("BestTester.xlsx")

#######라벨 생성#######
#메인 라벨
mainLabelFrame=LabelFrame(window, text="우수 테스터 선발")
mainLabelFrame.place(relx=0.01,relwidth=0.98,relheight=0.99)

#서브 라벨1
subLabelFrame1=LabelFrame(mainLabelFrame, text="Setting")
subLabelFrame1.place(relx=0.01, relwidth=0.98, relheight=0.50)
#서브 라벨2
subLabelFrame2=LabelFrame(mainLabelFrame, text="Result")
subLabelFrame2.place(relx=0.01, rely=0.51, relwidth=0.98, relheight=0.48)

#######버튼 생성#######
#파일 선택 버튼
fileopenBtn=tkinter.Button(subLabelFrame1,text="1) 유효 버그리포트 추출",width="30",command=fileOpen)
fileopenBtn.place(relx=0.1,rely=0.1)

fileEditBtn=tkinter.Button(subLabelFrame1,text="2) 우수테스터 선발하기", width="30", command=fileEdit)
fileEditBtn.place(relx=0.1, rely=0.6)

#######result 라벨 출력#######
#result 라벨 표시
log_text=tkinter.Label(subLabelFrame2,wraplength=280,justify="left", bg="blue")
log_text.place(relx=0.01,rely=0.01, relwidth=0.98, relheight=0.98)

window.mainloop()