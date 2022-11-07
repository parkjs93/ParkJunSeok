import tkinter
import tkinter.ttk
from tkinter import filedialog
from openpyxl import load_workbook
import pathlib
import os


###윈도우 생성하기###
window = tkinter.Tk()
window.title("스크립트 맞춤법 검사기")
window.geometry("350x400+200+200")
window.wm_iconbitmap('babyyeti.ico') #아이콘 생성
window.resizable(False,False)

###노트북(바탕) 만들기###
notebook = tkinter.ttk.Notebook(window)
notebook.place(relx=0.01,relwidth=0.98, relheight=0.91)

#UI 부분
frame_main = tkinter.Frame(window)
notebook.add(frame_main, text="Main")
frame_sentence = tkinter.Frame(window)
frame_item = tkinter.Frame(window)

#####함수 부분####
#프레임 지워주기
def frame_clear():
    if(notebook.tabs() == ('.!frame',)):
        notebook.forget(frame_main)
    elif(notebook.tabs() == ('.!frame2',)):
        notebook.forget(frame_sentence)
    elif(notebook.tabs() == ('.!frame3',)):
        notebook.forget(frame_item)
    else:
        pass

#프레임 리셋
def frame_reset():
    frame_clear()
    notebook.add(frame_main, text="Main")

#메뉴 선택 후 프레임 출력
def frame_active(framename):
    ####국내####
    if(framename == "sentence"):
        frame_clear()
        notebook.add(frame_sentence, text="Script")
    elif(framename == "item"):
        frame_clear()
        notebook.add(frame_item, text="Img File")
    else:
        frame_clear()
        notebook.add(frame_main, text="Main")
    
#대사 부분만 출력하기
def extraction(text):
    #대사 함수 추출해서 저장하기
    for fs in range(0,len(sayfunc)):
        find_funcIndex = text.find(sayfunc[fs])
        #sayfunc = 미리 선언한 대사 함수 유형 리스트에 포함될 경우 (ex. 함수명("유저","대사"))
        if(find_funcIndex != -1):
            fun_sentence = text[find_funcIndex:]
            fun_sentence = fun_sentence[:fun_sentence.find("(")]
            saySentence_Index = text.find('"')
            if (saySentence_Index == -1):
                pass
            else:
                saySentence = text[saySentence_Index+1:]
                saySentence_Index = saySentence.find('"')
                saySentence = saySentence[saySentence_Index+1:]
                saySentence_Index = saySentence.find('"')
                saySentence = saySentence[saySentence_Index+1:]
                saySentence_Index = saySentence.find('"')
                saySentence = saySentence[:saySentence_Index]
                return saySentence

        #sayScreenc = 미리 선언한 대사 함수 유형에 포함되지 않을 경우 다른 유형에 포함되어있는지 확인
        elif (find_funcIndex == -1):
            for fs2 in range(0, len(sayScreenc)):
                find_funcIndex = text.find(sayScreenc[fs2])
                if (find_funcIndex != -1):
                    fun_sentence = text[find_funcIndex:]
                    fun_sentence =fun_sentence[:fun_sentence.find("(")]

                    saySentence_Index = text.find('"')
                    if(saySentence_Index == -1):
                        pass
                    else:
                        saySentence = text[saySentence_Index+1:]
                        saySentence_Index = saySentence.find('"')
                        saySentence = saySentence[:saySentence_Index]
                        return saySentence

#파일 오픈 함수 (대사 오타)
def txtFileOpen(testFile, codeIdList):
    #파일 탐색기 오픈
    txtfilename = filedialog.askopenfilename(initialdir="",title="Select File")
    if(txtfilename == ""):
        #파일 탐색기 오픈 후 취소 할 경우 동작하지 않도록 하기 위해
        pass
    elif(txtfilename != ""):
        ######엑셀 작업부분######
        #지정된 파일 불러오기 (뒤 인자값 xlsm 데이터 적용을 위함) !!!!!지금 테스트를 위해 T 파일로 해뒀으니 나중에 변경할 것!!!!!
        wb=load_workbook("ScriptCheck.xlsm",read_only=False,keep_vba=True)#VBA 매크로 형태로 파일 오픈할 때 선언
        ws=wb['script']
        #엑셀에 값이 있는지 확인 후 빈 값 아니면 지우기
        row_check=5
        for r in ws.rows:
            for c in range(6):
                if(ws.cell(row=row_check, column=2+c).value == None):
                    pass
                else:
                    ws.cell(row=row_check, column=2+c).value = None
            row_check += 1
        #파일 오픈하여 lines 리스트 변수에 저장
        path = pathlib.Path(txtfilename)
        with open(path,"r",encoding="utf8") as file: #utf16은 .s파일 오픈할 때만... .txt
            try:
                lines = file.readlines()          #파일의 모든 줄을 읽어 lines 리스트에 저장
            except UnicodeError:                    #UnicodeError 발생 시 utf8로 인코딩 진행
                with open(path,"r",encoding="utf16") as file:
                    lines = file.readlines()
        #스크립트
        if(testFile=="script"):
            #한 줄 씩 엑셀에 값 입력하기
            j=5
            for line in lines:
                scriptSentence = extraction(line)   #대사만 추출해내는 함수 동작
                if(scriptSentence != None):
                    ws.cell(row=j, column=2).value = scriptSentence
                    j+=1
                else:
                    pass
        #아이템
        elif(testFile == "img"):
            itemInfoLlst =[]
            #변환된 IMG의 내용이 저장되며 문자열 특정 단어 치환
            for ls in range(0, len(lines)):
                a = lines[ls].replace("<tr><td>","")
                a = a.replace("</td><td>","//")
                a = a.replace("</td></tr>","")
                a = a.replace("\n"," ")

                for code in range(0, len(codeIdList)):
                    if(a.find(str(codeIdList[code])) != -1):
                        nameIndex = a.find("//")
                        itemName=a[:nameIndex]
                        a_slice=a[nameIndex+1:]

                        descIndex = a_slice.find("//")
                        itemDesc=a_slice[descIndex+2:]
                        itemInfoLlst.append("["+itemName+"] "+itemDesc)
                else:
                    pass
            for itemInfo in range(0,len(itemInfoLlst)):
                ws.cell(row=5+itemInfo, column=2).value = itemInfoLlst[itemInfo]
        else:
            print("error : "+str(testFile))
        #파일 저장
        wb.save("ScriptCheck.xlsm")
        os.system("start ScriptCheck.xlsm")  #노트북에선 주석처리
            
        print("end")

#줄바꿈 기준으로 값 저장시키기
def wordList(idcode):
    itemcode = []
    for id in range(0, len(idcode)):
        itemcode_index = idcode.find("\n")
        itemcode.append(str(idcode[:itemcode_index]))
        idcode = idcode[itemcode_index+1:]
        if(len(idcode)==0):
            #리스트 중 빈 값이 있으면 모든 값 출력되고있어 마지막에 빈 값 제거
            for nullId in range (0,len(itemcode)):
                if(itemcode[nullId]==""):
                    itemcode.remove("")
                else:
                    pass
            return itemcode
    

#Text 박스에서 입력받은 값 리스트에 저장시키기
def codeInsert(): 
    codeEntry = korItem_Entry.get(1.0, "end") 
    codeIdList=wordList(codeEntry)
    print(codeIdList)
    txtFileOpen("img",codeIdList)

def help():
    os.system("start https://confluence.nexon.com/pages/viewpage.action?pageId=551045126")

######비슷한 구조의 대사함수 리스트화 하기######
# "NPC","대사" : 두번째 인자에 대사 나오는 함수
sayfunc = ['sayface2','sayface', 'askYesNoface','adventuretpcheck','askface','askfaceMenu','sayTimeface','popface']
# "대사" : 바로 대사 나오는 함수
sayScreenc= ['speechBalloonEffect','inGameMonologue','askYesNoNpc','askText','target.messageFont','target.message','sayNpcNoESC','sayNpc','sayUserNoESC','sayUser','textEffectAd','textEffect','askMenuNpcNoESC']

###메뉴바 추가###
menubar = tkinter.Menu(window)

menu_kor = tkinter.Menu(menubar, tearoff=0)
menu_kor.add_command(label="Script", command=lambda:frame_active("sentence"))
menu_kor.add_command(label="Img File", command=lambda:frame_active("item"))
menubar.add_cascade(label="[맞춤법 및 번역 검사]", menu=menu_kor)
menu_kor.add_separator()
menu_kor.add_cascade(label="Help", command=help)
menu_kor.add_cascade(label="Main", command=frame_reset)

window.config(menu=menubar)

###페이지 라벨 만들기###
#메인 라벨
mainFrame = tkinter.LabelFrame(frame_main)
mainFrame.place(relx=0.01, relwidth=0.98, relheight=0.99)
nameLabel=tkinter.Label(window, text=" @hjong  @jeonyul  @parkjs", bg="black",fg="white")
nameLabel.place(relx=0.5,rely=0.93)

main_textLable = tkinter.Label(frame_main, text="맞춤법 검사 및 미번역 체크 관리 툴!\n\n\n상단의 메뉴에서 테스트 항목을 선택해주세요!", justify="left")
main_textLable.place(relx=0.05,rely=0.35)

#대사 맞춤법 검사 라벨
korScriptLF=tkinter.LabelFrame(frame_sentence)
korScriptLF.place(relx=0.01,relwidth=0.98,relheight=0.99)
korScript_txt=tkinter.Label(korScriptLF, text="원하는 대사 스크립트 파일을 선택해주세요.", justify="left")
korScript_txt.place(relx=0.1, rely=0.3)
korScript_txt=tkinter.Label(korScriptLF, text="경로 : ACData_KR\DataSvr\Script", justify="left")
korScript_txt.place(relx=0.1, rely=0.45)
korfileOpenBtn = tkinter.Button(korScriptLF, width=25, text="스크립트 파일 불러오기", command=lambda:txtFileOpen("script",""))
korfileOpenBtn.place(relx=0.2, rely=0.6)

#아이템 맞춤법 검사 라벨 
korItemLF=tkinter.LabelFrame(frame_item)
korItemLF.place(relx=0.01,relwidth=0.98,relheight=0.99)
korItem_Entry = tkinter.Text(korItemLF, width=30, height=23)
korItem_Entry.insert(1.0,"검색할 코드 입력해주세요.\n\n.img 파일을 .xls 파일로 변환 후 진행 부탁드립니다.\n\n(해당 부분 모두 지우고 입력)")
korItem_Entry.place(relx=0.01, rely=0.01)
korItem_EntryBtn = tkinter.Button(korItemLF, width=10, text="파일 선택", command=codeInsert)
korItem_EntryBtn.place(relx=0.7, rely=0.02)

window.mainloop()
